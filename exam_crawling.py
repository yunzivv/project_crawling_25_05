import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import time
import os
import re
import requests

# ocr-env_exam\Scripts\activate
# python exam_crawling.py

# # 게시판 링크 가져오기
# # 웹 페이지 요청
# url = 'https://www.comcbt.com/xe/anne'
# response = requests.get(url)
# soup = BeautifulSoup(response.text, 'html.parser')

# # CSS 선택자에 맞는 링크들 추출
# links = soup.select('#gnb > ul > li:nth-child(2) > ul > li:nth-child(6) > ul li > a')

# # 딕셔너리 형태로 파싱
# data = []
# for link in links:
#     title = link.get_text(strip=True)
#     href = link.get('href')
#     full_url = requests.compat.urljoin(url, href)  # 상대 경로 보완
#     data.append({'자격등급': '산업기사', 
#                  '자격증명': title, 
#                  'href': full_url, 
#                  '종류': '필기', 
#                  'regDate': time.strftime('%Y.%m.%d - %H:%M:%S')})

# file_path = 'exam.xlsx'

# # 데이터프레임 생성
# df_new = pd.DataFrame(data)

# # 기존 파일의 첫 번째 시트의 마지막 행 다음부터 이어쓰기
# book = load_workbook(file_path)
# sheetname = book.sheetnames[0]
# start_row = book[sheetname].max_row

# # 기존 파일에 이어쓰기
# with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     df_new.to_excel(writer, sheet_name=sheetname, startrow=start_row, index=False, header=False)


# 게시판에서 hwp 다운로드

BASE_URL = "https://www.comcbt.com"
SAVE_DIR = "hwp_files"
os.makedirs(SAVE_DIR, exist_ok=True)

def get_post_links(board_url):
    try:
        res = requests.get(board_url)
        res.raise_for_status()
    except Exception as e:
        print(f"게시판 페이지 요청 실패: {board_url} - {e}")
        return []
    
    soup = BeautifulSoup(res.text, 'html.parser')
    posts = []
    for a in soup.select('td.title a[href]'):
        title = a.get_text(strip=True)
        
        # (복원중) 포함 시 건너뛰기
        if '(복원중)' in title:
            print(f"⏭️ 건너뜀 (복원중): {title}")
            continue

        # 제목에서 연도 추출 (4자리 숫자)
        year_match = re.search(r'(20\d{2})', title)
        if year_match:
            year = int(year_match.group(1))
            if year < 2020:
                print(f"⏭️ 건너뜀 (연도미달): {title}")
                continue
        else:
            print(f"⏭️ 건너뜀 (연도없음): {title}")
            continue
        
        href = a['href']
        full_url = requests.compat.urljoin(BASE_URL, href)
        posts.append((title, full_url))
    return posts

def download_hwp_from_post(title, post_url):
    print(f"➡️ 게시글 접속: {title} - {post_url}")
    try:
        res = requests.get(post_url)
        res.raise_for_status()
    except Exception as e:
        print(f"게시글 페이지 요청 실패: {post_url} - {e}")
        return
    
    soup = BeautifulSoup(res.text, 'html.parser')
    links = soup.find_all('a')

    hwp_links = []
    for link in links:
        link_text = link.get_text(strip=True)
        if link_text.endswith('.hwp') and '(교사용)' in link_text:
            href = link.get('href')
            if href:
                hwp_links.append((link_text, href))

    if not hwp_links:
        print("  ⚠️ '(교사용).hwp' 링크 없음")
        return

    link_text, href = hwp_links[0]
    file_url = requests.compat.urljoin(BASE_URL, href)
    # () 앞까지 잘라서 파일명 생성
    filename = link_text.split('(')[0].strip() + '.hwp'
    save_path = os.path.join(SAVE_DIR, filename)

    if os.path.exists(save_path):
        print(f"  이미 존재하는 파일: {filename} (다운로드 건너뜀)")
        return

    print(f"📥 다운로드 중: {link_text} - 파일명: {filename}")
    try:
        file_content = requests.get(file_url).content
        with open(save_path, 'wb') as f:
            f.write(file_content)
    except Exception as e:
        print(f"  다운로드 실패: {file_url} - {e}")

def main():
    # 엑셀에서 게시판 URL 리스트 읽기
    df = pd.read_excel('exam_board.xlsx')
    board_urls = df['href'].tolist()

    for board_url in board_urls:
        full_board_url = requests.compat.urljoin(BASE_URL, board_url)
        print(f"\n▶ 게시판 접속: {full_board_url}")
        posts = get_post_links(full_board_url)
        print(f"게시글 개수: {len(posts)}")

        for title, post_url in posts:
            download_hwp_from_post(title, post_url)

if __name__ == '__main__':
    main()
